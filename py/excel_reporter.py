"""
엑셀 리포터 모듈
추천 점수 순 1~30위를 엑셀 파일로 출력합니다.
"""

import os
import re
import json
from datetime import datetime
from email.utils import parsedate_to_datetime
from typing import Optional

import pandas as pd


def _normalize_date(value) -> str:
    """업로드일을 YYYY-MM-DD 형식으로 정규화."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    try:
        if "T" in s:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        elif re.match(r"^\d{4}-\d{2}-\d{2}", s):
            dt = datetime.strptime(s[:10], "%Y-%m-%d")
        else:
            dt = parsedate_to_datetime(s)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""


# 출력 컬럼 순서 (순위는 정렬 후 추가)
OUTPUT_COLUMNS_BASE = [
    "제목",
    "추천점수",
    "키워드",
    "카테고리",
    "출처",
    "유튜브_URL",
    "뉴스기사_URL",
    "업로드일",
    "뉴스기사2_URL",
    "뉴스기사2_날짜",
    "뉴스기사3_URL",
    "뉴스기사3_날짜",
    "조회수",
]
OUTPUT_COLUMNS = ["순위"] + OUTPUT_COLUMNS_BASE

# 컬럼별 최소 너비 (한글 가독성)
COLUMN_WIDTHS = {
    "순위": 8,
    "제목": 50,
    "추천점수": 12,
    "키워드": 14,
    "카테고리": 10,
    "출처": 14,
    "유튜브_URL": 50,
    "뉴스기사_URL": 50,
    "업로드일": 14,
    "뉴스기사2_URL": 50,
    "뉴스기사2_날짜": 14,
    "뉴스기사3_URL": 50,
    "뉴스기사3_날짜": 14,
    "조회수": 14,
}

TOP_N = 30


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """필수 컬럼이 없으면 추가 (빈 값으로)."""
    column_mapping = {
        "title": "제목",
        "score": "추천점수",
        "score_keywords": "키워드",
        "category": "카테고리",
        "source": "출처",
        "youtube_url": "유튜브_URL",
        "news_url": "뉴스기사_URL",
        "url": "뉴스기사_URL",  # 스크래퍼에서 url로 넘길 때
        "views": "조회수",
        "upload_date": "업로드일",
    }
    out = df.copy()
    for eng, kor in column_mapping.items():
        if kor not in out.columns and eng in out.columns:
            out[kor] = out[eng]
    for col in OUTPUT_COLUMNS_BASE:
        if col not in out.columns:
            out[col] = ""
    return out[OUTPUT_COLUMNS_BASE]


def _set_column_widths(worksheet) -> None:
    """엑셀 컬럼 너비 설정."""
    try:
        from openpyxl.utils import get_column_letter

        for idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            width = COLUMN_WIDTHS.get(col_name, 15)
            worksheet.column_dimensions[get_column_letter(idx)].width = width
    except Exception as e:
        print(f"[경고] 컬럼 너비 설정 중 오류: {e}")


# 가운데 정렬할 컬럼 (순위, 추천점수, 키워드, 출처, 카테고리)
CENTER_ALIGN_COLUMNS = {"순위", "추천점수", "키워드", "출처", "카테고리"}


def _set_header_center_alignment(worksheet, num_columns: int) -> None:
    """제목라인(헤더 1행) 전체 가운데 정렬."""
    try:
        from openpyxl.styles import Alignment

        center = Alignment(horizontal="center", vertical="center")
        for col_idx in range(1, num_columns + 1):
            worksheet.cell(row=1, column=col_idx).alignment = center
    except Exception as e:
        print(f"[경고] 헤더 가운데 정렬 설정 중 오류: {e}")


def _set_center_alignment_columns(worksheet, num_data_rows: int) -> None:
    """순위·추천점수·키워드·출처 컬럼 가운데 정렬 (데이터 행)."""
    try:
        from openpyxl.styles import Alignment

        center = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            if col_name not in CENTER_ALIGN_COLUMNS:
                continue
            for row in range(2, num_data_rows + 2):  # 데이터 행만 (헤더 제외)
                cell = worksheet.cell(row=row, column=col_idx)
                cell.alignment = center
    except Exception as e:
        print(f"[경고] 가운데 정렬 설정 중 오류: {e}")


def _apply_hyperlink_style(cell, url: str) -> None:
    """셀에 하이퍼링크 및 파란색 밑줄 스타일 적용."""
    if not url or not isinstance(url, str) or not url.strip().startswith("http"):
        return
    try:
        from openpyxl.styles import Font

        cell.hyperlink = url.strip()
        cell.font = Font(color="0563C1", underline="single")
    except Exception:
        pass


def _apply_hyperlinks(worksheet, df: pd.DataFrame) -> None:
    """유튜브_URL, 뉴스기사_URL(1~3) 컬럼에 하이퍼링크 적용."""
    try:
        url_columns = ["유튜브_URL", "뉴스기사_URL", "뉴스기사2_URL", "뉴스기사3_URL"]
        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2  # 헤더 1행 + 0-based
            for col_name in url_columns:
                if col_name not in OUTPUT_COLUMNS:
                    continue
                url = row.get(col_name, "") or ""
                if url and str(url).strip().startswith("http"):
                    col_idx = OUTPUT_COLUMNS.index(col_name) + 1
                    _apply_hyperlink_style(worksheet.cell(row=excel_row, column=col_idx), str(url))
    except Exception as e:
        print(f"[경고] 하이퍼링크 적용 중 오류: {e}")


def export_to_excel(
    df: pd.DataFrame,
    output_path: Optional[str] = None,
    score_column: str = "추천점수",
    ascending: bool = False,
) -> str:
    """
    DataFrame을 엑셀 파일로 저장합니다.

    Args:
        df: 수집·분석된 데이터 (제목, 추천점수, 출처, URL 등 포함)
        output_path: 저장 경로. None이면 자동 생성 (agro_report_MMDD(1).xlsx, (2).xlsx, ...)
        score_column: 정렬에 사용할 점수 컬럼명
        ascending: False=높은순, True=낮은순

    Returns:
        저장된 파일 경로
    """
    try:
        if df.empty:
            raise ValueError("데이터가 비어 있습니다.")

        # 점수 컬럼 통일
        if "추천점수" not in df.columns and score_column in df.columns:
            df = df.copy()
            df["추천점수"] = df[score_column]

        # 컬럼 정규화
        out = _ensure_columns(df)

        # 추천점수 기준 정렬 (숫자 변환 시도)
        try:
            out["추천점수"] = pd.to_numeric(out["추천점수"], errors="coerce").fillna(0)
        except Exception:
            pass
        out = out.sort_values(by="추천점수", ascending=ascending)

        # 1~30위만 선택
        out = out.head(TOP_N).reset_index(drop=True)
        out.insert(0, "순위", list(range(1, len(out) + 1)))

        # 업로드일·뉴스기사 날짜 YYYY-MM-DD 형식으로 정규화
        for col in ("업로드일", "뉴스기사2_날짜", "뉴스기사3_날짜"):
            if col in out.columns:
                out[col] = out[col].apply(_normalize_date)

        # 출력 경로 결정: xlsx/ 폴더에 agro_report_MMDD(1).xlsx, (2).xlsx, ... 순서
        if not output_path:
            # 프로젝트 루트 기준 xlsx 폴더
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            output_dir = os.path.join(base_dir, "xlsx")
            
            os.makedirs(output_dir, exist_ok=True)
            base = f"agro_report_{datetime.now().strftime('%m%d')}"
            n = 1
            output_path = os.path.join(output_dir, f"{base}({n}).xlsx")
            while os.path.exists(output_path):
                n += 1
                output_path = os.path.join(output_dir, f"{base}({n}).xlsx")

        output_path = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # 엑셀 저장 (openpyxl 엔진으로 컬럼 너비·하이퍼링크 조정)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="어그로추천주제")
            worksheet = writer.sheets["어그로추천주제"]
            _set_column_widths(worksheet)
            _set_header_center_alignment(worksheet, len(OUTPUT_COLUMNS))
            _set_center_alignment_columns(worksheet, len(out))
            _apply_hyperlinks(worksheet, out)

        return output_path

        return output_path

    except ValueError as e:
        raise e
    except Exception as e:
        raise RuntimeError(f"엑셀 저장 실패: {e}") from e


def export_to_json(
    df: pd.DataFrame,
    output_path: Optional[str] = None,
    score_column: str = "추천점수",
    ascending: bool = False,
) -> str:
    """
    DataFrame을 웹용 JSON 파일로 저장합니다.
    """
    try:
        if df.empty:
            return ""

        # 점수 컬럼 통일
        if "추천점수" not in df.columns and score_column in df.columns:
            df = df.copy()
            df["추천점수"] = df[score_column]

        # 컬럼 정규화
        out = _ensure_columns(df)

        # 추천점수 기준 정렬
        try:
            out["추천점수"] = pd.to_numeric(out["추천점수"], errors="coerce").fillna(0)
        except Exception:
            pass
        out = out.sort_values(by="추천점수", ascending=ascending)

        # 1~30위만 선택
        out = out.head(TOP_N).reset_index(drop=True)
        out.insert(0, "순위", list(range(1, len(out) + 1)))

        # 날짜 정규화
        for col in ("업로드일", "뉴스기사2_날짜", "뉴스기사3_날짜"):
            if col in out.columns:
                out[col] = out[col].apply(_normalize_date)

        # 저장 경로: web/data.json
        if not output_path:
            output_dir = "web"
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, "data.json")

        output_path = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # JSON 저장
        data = out.to_dict(orient="records")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return output_path

    except Exception as e:
        print(f"[오류] JSON 저장 실패: {e}")
        return ""


def export_to_js(
    df: pd.DataFrame,
    output_path: Optional[str] = None,
    score_column: str = "추천점수",
    ascending: bool = False,
    scraper_status: dict = None,
) -> str:
    """
    DataFrame을 웹용 JS 파일로 저장합니다.
    const keywordData = [...]; 형태로 저장되어
    HTML에서 <script src="data.js"></script>로 불러올 수 있습니다.
    """
    try:
        if df.empty:
            return ""

        # 점수 컬럼 통일
        if "추천점수" not in df.columns and score_column in df.columns:
            df = df.copy()
            df["추천점수"] = df[score_column]

        # 컬럼 정규화
        out = _ensure_columns(df)

        # 추천점수 기준 정렬
        try:
            out["추천점수"] = pd.to_numeric(out["추천점수"], errors="coerce").fillna(0)
        except Exception:
            pass
        out = out.sort_values(by="추천점수", ascending=ascending)

        # 1~30위만 선택
        out = out.head(TOP_N).reset_index(drop=True)
        out.insert(0, "순위", list(range(1, len(out) + 1)))

        # 날짜 정규화
        for col in ("업로드일", "뉴스기사2_날짜", "뉴스기사3_날짜"):
            if col in out.columns:
                out[col] = out[col].apply(_normalize_date)

        # 저장 경로: ../data.js (루트)
        if not output_path:
            # 프로젝트 루트 기준 data.js
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            output_path = os.path.join(base_dir, "data.js")

        output_path = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # JS 저장
        data = out.to_dict(orient="records")
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        
        status_str = json.dumps(scraper_status or {}, ensure_ascii=False, indent=2)
        
        js_content = f"const keywordData = {json_str};\n"
        js_content += f"const scraperStatus = {status_str};\n"

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(js_content)

        return output_path

    except Exception as e:
        print(f"[오류] JS 저장 실패: {e}")
        return ""
